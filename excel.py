import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from datetime import datetime, timedelta

# 从CSV文件中读取项目数据
df = pd.read_csv('projects.csv')

# 将日期转换为datetime对象
df['开始时间'] = pd.to_datetime(df['开始时间'], format='%Y%m%d')
df['结束时间'] = pd.to_datetime(df['结束时间'], format='%Y%m%d')

# 填充模块的空白值
df['模块'] = df['模块'].ffill()
# 填充空白的开始时间
for i in range(1, len(df)):
    if pd.isna(df.at[i, '开始时间']):
        previous_end_date = df.at[i-1, '结束时间']
        new_start_date = previous_end_date + timedelta(days=1)
        df.at[i, '开始时间'] = new_start_date

# 创建一个新的Excel工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Gantt Chart"

# 设置表头
headers = ['关键里程碑', '开始时间', '结束时间', '负责人', '模块']
ws.append(headers)

# 添加数据到表中
for index, row in df.iterrows():
    ws.append([row['关键里程碑'], row['开始时间'].strftime('%Y-%m-%d'), row['结束时间'].strftime('%Y-%m-%d'), row['负责人'], row['模块']])

# 获取开始和结束日期的范围
start_date = df['开始时间'].min()
end_date = df['结束时间'].max()
date_range = pd.date_range(start=start_date, end=end_date)

# 添加日期作为列
for i, date in enumerate(date_range):
    ws.cell(row=1, column=len(headers) + i + 1, value=date.strftime('%Y-%m-%d'))

# 填充甘特图
color_mapping = {
    '工艺': 'FFC7CE',  # 红色
    '数采': 'FFEB9C',  # 黄色
    '算法': 'C6EFCE',  # 绿色
    '电控': 'D9EAD3'   # 紫色
}

for i, row in df.iterrows():
    start_col = (row['开始时间'] - start_date).days + len(headers) + 1
    end_col = (row['结束时间'] - start_date).days + len(headers) + 1
    fill_color = color_mapping.get(row['负责人'], 'FFFFFF')  # 默认填充白色

    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=i + 2, column=col)
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

# 保存为Excel文件
wb.save("gantt_chart.xlsx")
